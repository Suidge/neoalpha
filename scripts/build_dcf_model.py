#!/usr/bin/env python3
"""
Build a standalone DCF workbook with live Excel formulas.

This is a lightweight OpenClaw adaptation of the Excel modeling conventions from
anthropics/financial-services (Apache 2.0). It intentionally writes formulas,
not precomputed valuation outputs, so assumptions remain editable in Excel.
"""

from __future__ import annotations

import argparse
import json
from datetime import date
from pathlib import Path
from typing import Any, Dict, List

try:
    from openpyxl import Workbook
    from openpyxl.comments import Comment
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError as exc:  # pragma: no cover - dependency gate
    raise SystemExit(
        "openpyxl is required. Install with: python3 -m pip install openpyxl"
    ) from exc

BLUE_FONT = "0000FF"
BLACK_FONT = "000000"
GREEN_FONT = "008000"
WHITE_FONT = "FFFFFF"
DARK_BLUE = "1F4E79"
LIGHT_BLUE = "D9E1F2"
LIGHT_GREY = "F2F2F2"
BASE_BLUE = "BDD7EE"
THIN_GREY = "BFBFBF"


def pct(x: Any) -> float:
    if x is None:
        return 0.0
    if isinstance(x, str):
        s = x.strip().replace("%", "")
        v = float(s)
        return v / 100 if "%" in x or v > 1 else v
    return float(x)


def money(x: Any) -> float:
    if x is None:
        return 0.0
    if isinstance(x, str):
        return float(x.replace(",", ""))
    return float(x)


def default_payload() -> Dict[str, Any]:
    return {
        "company": "Example Company",
        "ticker": "EXAMPLE.US",
        "currency": "USD",
        "units": "millions",
        "as_of": date.today().isoformat(),
        "source": "User provided / Longbridge verified data",
        "historical": {
            "revenue": 1000,
            "ebit_margin": 0.18,
            "tax_rate": 0.21,
            "da_pct_revenue": 0.04,
            "capex_pct_revenue": 0.05,
            "nwc_pct_revenue_change": 0.01,
        },
        "market": {
            "cash": 100,
            "debt": 250,
            "minority_interest": 0,
            "preferred_equity": 0,
            "non_operating_assets": 0,
            "diluted_shares": 100,
            "current_price": 12,
        },
        "wacc": {
            "risk_free_rate": 0.045,
            "beta": 1.10,
            "equity_risk_premium": 0.055,
            "size_country_premium": 0.0,
            "pre_tax_cost_of_debt": 0.06,
            "debt_weight": 0.20,
        },
        "scenarios": {
            "Bear": {"revenue_growth": 0.06, "ebit_margin": 0.15, "terminal_growth": 0.02},
            "Base": {"revenue_growth": 0.10, "ebit_margin": 0.18, "terminal_growth": 0.03},
            "Bull": {"revenue_growth": 0.14, "ebit_margin": 0.22, "terminal_growth": 0.035},
        },
    }


def load_payload(path: str | None) -> Dict[str, Any]:
    payload = default_payload()
    if not path:
        return payload
    user = json.loads(Path(path).read_text())
    # shallow/deep merge for known nested dicts
    for k, v in user.items():
        if isinstance(v, dict) and isinstance(payload.get(k), dict):
            payload[k].update(v)
        else:
            payload[k] = v
    return payload


def style_input(cell, source: str):
    cell.font = Font(color=BLUE_FONT)
    cell.fill = PatternFill("solid", fgColor=LIGHT_GREY)
    cell.comment = Comment(f"Source: {source}", "OpenClaw")


def style_formula(cell):
    cell.font = Font(color=BLACK_FONT)


def section(ws, row: int, title: str, start_col: int = 1, end_col: int = 9):
    ws.cell(row, start_col).value = title
    ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=end_col)
    c = ws.cell(row, start_col)
    c.fill = PatternFill("solid", fgColor=DARK_BLUE)
    c.font = Font(color=WHITE_FONT, bold=True)
    c.alignment = Alignment(horizontal="center")


def header_row(ws, row: int, cols: int):
    for col in range(1, cols + 1):
        c = ws.cell(row, col)
        c.fill = PatternFill("solid", fgColor=LIGHT_BLUE)
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center")


def apply_grid(ws):
    side = Side(style="thin", color=THIN_GREY)
    for row in ws.iter_rows():
        for c in row:
            if c.value is not None:
                c.border = Border(left=side, right=side, top=side, bottom=side)
                if isinstance(c.value, str) and c.value.startswith("="):
                    style_formula(c)


def build_workbook(payload: Dict[str, Any], output: Path, years: int = 5) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Inputs"
    dcf = wb.create_sheet("DCF")
    wacc = wb.create_sheet("WACC")
    sens = wb.create_sheet("Sensitivity")
    checks = wb.create_sheet("Checks")

    source = f"{payload.get('source', 'User provided')}, {payload.get('as_of', date.today().isoformat())}"
    scenarios = payload["scenarios"]
    hist = payload["historical"]
    market = payload["market"]
    w = payload["wacc"]

    # Inputs
    section(ws, 1, f"{payload['ticker']} — {payload['company']} DCF Inputs")
    rows = [
        ("Company", payload["company"]),
        ("Ticker", payload["ticker"]),
        ("Currency", payload.get("currency", "USD")),
        ("Units", payload.get("units", "millions")),
        ("As of", payload.get("as_of", date.today().isoformat())),
        ("Scenario selector (1 Bear / 2 Base / 3 Bull)", 2),
        ("Latest revenue", money(hist["revenue"])),
        ("Tax rate", pct(hist["tax_rate"])),
        ("D&A % revenue", pct(hist["da_pct_revenue"])),
        ("CapEx % revenue", pct(hist["capex_pct_revenue"])),
        ("NWC % revenue change", pct(hist["nwc_pct_revenue_change"])),
        ("Cash", money(market["cash"])),
        ("Debt", money(market["debt"])),
        ("Minority interest", money(market.get("minority_interest", 0))),
        ("Preferred equity", money(market.get("preferred_equity", 0))),
        ("Non-operating assets", money(market.get("non_operating_assets", 0))),
        ("Diluted shares", money(market["diluted_shares"])),
        ("Current price", money(market["current_price"])),
    ]
    ws["A3"], ws["B3"] = "Input", "Value"
    header_row(ws, 3, 2)
    input_row = {}
    for idx, (label, value) in enumerate(rows, start=4):
        ws.cell(idx, 1).value = label
        ws.cell(idx, 2).value = value
        input_row[label] = idx
        style_input(ws.cell(idx, 2), source)

    scen_start = 24
    section(ws, scen_start - 1, "Scenario Assumptions", 1, 5)
    ws.cell(scen_start, 1).value = "Scenario"
    ws.cell(scen_start, 2).value = "Revenue growth"
    ws.cell(scen_start, 3).value = "EBIT margin"
    ws.cell(scen_start, 4).value = "Terminal growth"
    header_row(ws, scen_start, 4)
    scen_rows = {}
    for i, name in enumerate(["Bear", "Base", "Bull"], start=scen_start + 1):
        scen_rows[name] = i
        ws.cell(i, 1).value = name
        vals = scenarios[name]
        for col, key in enumerate(["revenue_growth", "ebit_margin", "terminal_growth"], start=2):
            ws.cell(i, col).value = pct(vals[key])
            style_input(ws.cell(i, col), source)

    # WACC
    section(wacc, 1, "WACC")
    wacc_rows = [
        ("Risk-free rate", pct(w["risk_free_rate"])),
        ("Beta", money(w["beta"])),
        ("Equity risk premium", pct(w["equity_risk_premium"])),
        ("Size/country premium", pct(w.get("size_country_premium", 0))),
        ("Cost of equity", "=B4+B5*B6+B7"),
        ("Pre-tax cost of debt", pct(w["pre_tax_cost_of_debt"])),
        ("Tax rate", f"=Inputs!B{input_row['Tax rate']}"),
        ("After-tax cost of debt", "=B9*(1-B10)"),
        ("Debt weight", pct(w["debt_weight"])),
        ("Equity weight", "=1-B12"),
        ("WACC", "=B13*B8+B12*B11"),
    ]
    wacc["A3"], wacc["B3"] = "Item", "Value"
    header_row(wacc, 3, 2)
    for r, (label, value) in enumerate(wacc_rows, start=4):
        wacc.cell(r, 1).value = label
        wacc.cell(r, 2).value = value
        if isinstance(value, (int, float)):
            style_input(wacc.cell(r, 2), source)

    # DCF
    section(dcf, 1, "DCF Projection")
    dcf["A3"] = "Line item"
    for i in range(years + 1):
        dcf.cell(3, 2 + i).value = "Actual" if i == 0 else f"Y{i}"
    header_row(dcf, 3, years + 2)
    base_rows = {
        "Revenue": 4,
        "Revenue growth": 5,
        "EBIT margin": 6,
        "EBIT": 7,
        "Tax": 8,
        "NOPAT": 9,
        "D&A": 10,
        "CapEx": 11,
        "ΔNWC": 12,
        "FCF": 13,
        "Discount factor": 14,
        "PV FCF": 15,
    }
    for label, row in base_rows.items():
        dcf.cell(row, 1).value = label
    dcf["B4"] = f"=Inputs!B{input_row['Latest revenue']}"
    dcf["B5"] = ""
    dcf["B6"] = f"=INDEX(Inputs!C{scen_rows['Bear']}:C{scen_rows['Bull']},Inputs!B{input_row['Scenario selector (1 Bear / 2 Base / 3 Bull)']})"
    dcf["B7"] = "=B4*B6"
    dcf["B8"] = f"=B7*Inputs!B{input_row['Tax rate']}"
    dcf["B9"] = "=B7-B8"
    dcf["B10"] = f"=B4*Inputs!B{input_row['D&A % revenue']}"
    dcf["B11"] = f"=-B4*Inputs!B{input_row['CapEx % revenue']}"
    dcf["B12"] = "=0"
    dcf["B13"] = "=B9+B10+B11-B12"
    dcf["B14"] = "=1"
    dcf["B15"] = "=0"
    for col in range(3, years + 3):
        prev = get_column_letter(col - 1)
        cur = get_column_letter(col)
        dcf.cell(4, col).value = f"={prev}4*(1+{cur}5)"
        dcf.cell(5, col).value = f"=INDEX(Inputs!B{scen_rows['Bear']}:B{scen_rows['Bull']},Inputs!B{input_row['Scenario selector (1 Bear / 2 Base / 3 Bull)']})"
        dcf.cell(6, col).value = f"=INDEX(Inputs!C{scen_rows['Bear']}:C{scen_rows['Bull']},Inputs!B{input_row['Scenario selector (1 Bear / 2 Base / 3 Bull)']})"
        dcf.cell(7, col).value = f"={cur}4*{cur}6"
        dcf.cell(8, col).value = f"={cur}7*Inputs!$B${input_row['Tax rate']}"
        dcf.cell(9, col).value = f"={cur}7-{cur}8"
        dcf.cell(10, col).value = f"={cur}4*Inputs!$B${input_row['D&A % revenue']}"
        dcf.cell(11, col).value = f"=-{cur}4*Inputs!$B${input_row['CapEx % revenue']}"
        dcf.cell(12, col).value = f"=({cur}4-{prev}4)*Inputs!$B${input_row['NWC % revenue change']}"
        dcf.cell(13, col).value = f"={cur}9+{cur}10+{cur}11-{cur}12"
        dcf.cell(14, col).value = f"=1/(1+WACC!$B$14)^{col-2}"
        dcf.cell(15, col).value = f"={cur}13*{cur}14"

    val_row = 18
    section(dcf, val_row - 1, "Valuation Summary", 1, 4)
    labels = [
        ("Terminal growth", f"=INDEX(Inputs!D{scen_rows['Bear']}:D{scen_rows['Bull']},Inputs!B{input_row['Scenario selector (1 Bear / 2 Base / 3 Bull)']})"),
        ("Terminal value", f"={get_column_letter(years+2)}13*(1+B{val_row})/(WACC!$B$14-B{val_row})"),
        ("PV terminal value", f"=B{val_row+1}*{get_column_letter(years+2)}14"),
        ("Enterprise value (EV)", f"=SUM(C15:{get_column_letter(years+2)}15)+B{val_row+2}"),
        ("Net debt", f"=Inputs!B{input_row['Debt']}-Inputs!B{input_row['Cash']}"),
        ("Minority interest", f"=Inputs!B{input_row['Minority interest']}"),
        ("Preferred equity", f"=Inputs!B{input_row['Preferred equity']}"),
        ("Non-operating assets", f"=Inputs!B{input_row['Non-operating assets']}"),
        ("Equity value", f"=B{val_row+3}-B{val_row+4}-B{val_row+5}-B{val_row+6}+B{val_row+7}"),
        ("Diluted shares", f"=Inputs!B{input_row['Diluted shares']}"),
        ("Intrinsic value / share", f"=B{val_row+8}/B{val_row+9}"),
        ("Current price", f"=Inputs!B{input_row['Current price']}"),
        ("Upside / downside", f"=B{val_row+10}/B{val_row+11}-1"),
    ]
    for i, (label, formula) in enumerate(labels, start=val_row):
        dcf.cell(i, 1).value = label
        dcf.cell(i, 2).value = formula

    # Sensitivity: WACC x terminal growth, full DCF formula using final FCF
    section(sens, 1, "Sensitivity — Intrinsic Value / Share")
    sens["A3"] = "WACC \\ Terminal g"
    base_wacc_cell = "WACC!$B$14"
    base_g_cell = f"DCF!$B${val_row}"
    terminal_fcf = f"DCF!${get_column_letter(years+2)}$13"
    pv_fcf_sum = f"SUM(DCF!$C$15:DCF!${get_column_letter(years+2)}$15)"
    net_debt = f"DCF!$B${val_row+4}"
    mi = f"DCF!$B${val_row+5}"
    pref = f"DCF!$B${val_row+6}"
    noa = f"DCF!$B${val_row+7}"
    shares = f"DCF!$B${val_row+9}"
    for j, delta_g in enumerate([-0.005, -0.0025, 0, 0.0025, 0.005], start=2):
        sens.cell(3, j).value = f"={base_g_cell}+{delta_g}"
    for i, delta_w in enumerate([-0.01, -0.005, 0, 0.005, 0.01], start=4):
        sens.cell(i, 1).value = f"={base_wacc_cell}+{delta_w}"
        for j in range(2, 7):
            wcell = f"$A{i}"
            gcell = f"{get_column_letter(j)}$3"
            tv = f"({terminal_fcf}*(1+{gcell})/({wcell}-{gcell}))"
            pvtv = f"({tv}/(1+{wcell})^{years})"
            ev = f"({pv_fcf_sum}+{pvtv})"
            sens.cell(i, j).value = f"=({ev}-{net_debt}-{mi}-{pref}+{noa})/{shares}"
    sens[6][2].fill = PatternFill("solid", fgColor=BASE_BLUE)  # C6 center cell
    sens[6][2].font = Font(bold=True)
    header_row(sens, 3, 6)

    # Checks
    section(checks, 1, "Model Checks")
    check_rows = [
        ("Terminal growth < WACC", f"=DCF!B{val_row}<WACC!B14"),
        ("Shares > 0", f"=DCF!B{val_row+9}>0"),
        ("Base sensitivity = Base value", f"=ABS(Sensitivity!D6-DCF!B{val_row+10})<0.01"),
        ("No negative WACC", "=WACC!B14>0"),
    ]
    checks["A3"], checks["B3"] = "Check", "Pass?"
    header_row(checks, 3, 2)
    for r, (label, formula) in enumerate(check_rows, start=4):
        checks.cell(r, 1).value = label
        checks.cell(r, 2).value = formula

    for sheet in wb.worksheets:
        for col in range(1, min(sheet.max_column, 12) + 1):
            sheet.column_dimensions[get_column_letter(col)].width = 18
        apply_grid(sheet)
        for row in sheet.iter_rows():
            for c in row:
                if isinstance(c.value, (int, float)) and c.row > 1:
                    c.number_format = '0.0%' if abs(c.value) < 1 else '#,##0.0'
                if isinstance(c.value, str) and c.value.startswith('='):
                    c.number_format = '0.0%'

    # Specific number formats after generic pass
    for sheet in [dcf, sens]:
        for row in sheet.iter_rows():
            for c in row:
                if c.value is not None:
                    if c.row in [5, 6, 14, val_row, val_row + 12] or sheet.title == "Sensitivity":
                        c.number_format = '0.0%'
                    else:
                        c.number_format = '#,##0.0'
    for r in range(val_row, val_row + len(labels)):
        dcf.cell(r, 2).number_format = '0.0%' if r in [val_row, val_row + 12] else '#,##0.0'
    dcf.cell(val_row + 10, 2).number_format = '#,##0.00'
    dcf.cell(val_row + 11, 2).number_format = '#,##0.00'

    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)


def main() -> None:
    parser = argparse.ArgumentParser(description="Build a formula-driven DCF workbook")
    parser.add_argument("--input", help="JSON input file. Omit for example model.")
    parser.add_argument("--output", default="out/model.xlsx", help="Output .xlsx path")
    parser.add_argument("--years", type=int, default=5, choices=range(3, 11), metavar="3-10")
    args = parser.parse_args()
    payload = load_payload(args.input)
    build_workbook(payload, Path(args.output), args.years)
    print(args.output)


if __name__ == "__main__":
    main()
